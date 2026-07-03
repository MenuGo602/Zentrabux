"""
Excel Export Service — hisobotlarni .xlsx formatida eksport qiladi.

Barcha eksport funksiyalari deterministik Engine natijalarini (Trial
Balance, P&L) yoki DB'dan olingan tranzaksiyalar ro'yxatini qabul qiladi —
o'zi hech qanday hisob-kitob qilmaydi.
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.engines.report.pnl import ProfitAndLoss
from app.engines.report.trial_balance import TrialBalance
from app.models.all_models import Transaction

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True)


class ExcelExportService:
    def export_transactions(self, transactions: list[Transaction], company_name: str = "") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tranzaksiyalar"

        headers = ["Sana", "Tavsif", "Turi", "Summa", "Valyuta", "Status", "AI orqali"]
        row = self._write_header(ws, headers, title=company_name)

        for tx in transactions:
            ws.cell(row=row, column=1, value=str(tx.transaction_date))
            ws.cell(row=row, column=2, value=tx.description)
            ws.cell(row=row, column=3, value=tx.transaction_type)
            ws.cell(row=row, column=4, value=float(tx.total_amount))
            ws.cell(row=row, column=5, value=tx.currency)
            ws.cell(row=row, column=6, value=tx.status)
            ws.cell(row=row, column=7, value="Ha" if tx.ai_generated else "Yo'q")
            row += 1

        total_row = row + 1
        ws.cell(row=total_row, column=3, value="Jami:").font = TOTAL_FONT
        total = sum((tx.total_amount for tx in transactions), Decimal("0"))
        ws.cell(row=total_row, column=4, value=float(total)).font = TOTAL_FONT

        self._auto_width(ws, [12, 45, 12, 16, 10, 12, 10])
        return self._to_bytes(wb)

    def export_trial_balance(self, trial_balance: TrialBalance, company_name: str = "") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Aylanma qaydnoma"

        headers = ["Kod", "Nomi", "Turi", "Debit", "Credit", "Qoldiq"]
        row = self._write_header(ws, headers, title=company_name)

        for line in trial_balance.lines:
            ws.cell(row=row, column=1, value=line.account_code)
            ws.cell(row=row, column=2, value=line.account_name)
            ws.cell(row=row, column=3, value=line.account_type)
            ws.cell(row=row, column=4, value=float(line.total_debit))
            ws.cell(row=row, column=5, value=float(line.total_credit))
            ws.cell(row=row, column=6, value=float(line.balance))
            row += 1

        total_row = row + 1
        ws.cell(row=total_row, column=2, value="Jami:").font = TOTAL_FONT
        ws.cell(row=total_row, column=4, value=float(trial_balance.total_debit)).font = TOTAL_FONT
        ws.cell(row=total_row, column=5, value=float(trial_balance.total_credit)).font = TOTAL_FONT

        balance_row = total_row + 1
        status = "✅ Balansda" if trial_balance.is_balanced else "❌ Balanssiz"
        ws.cell(row=balance_row, column=2, value=status).font = TOTAL_FONT

        self._auto_width(ws, [10, 35, 12, 16, 16, 16])
        return self._to_bytes(wb)

    def export_pnl(self, pnl: ProfitAndLoss, company_name: str = "") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Foyda va zarar"

        row = self._write_header(ws, ["Kod", "Nomi", "Summa"], title=company_name)

        ws.cell(row=row, column=1, value="DAROMADLAR").font = TOTAL_FONT
        row += 1
        for item in pnl.income_items:
            ws.cell(row=row, column=1, value=item.account_code)
            ws.cell(row=row, column=2, value=item.account_name)
            ws.cell(row=row, column=3, value=float(item.amount))
            row += 1
        ws.cell(row=row, column=2, value="Jami daromad:").font = TOTAL_FONT
        ws.cell(row=row, column=3, value=float(pnl.total_income)).font = TOTAL_FONT
        row += 2

        ws.cell(row=row, column=1, value="XARAJATLAR").font = TOTAL_FONT
        row += 1
        for item in pnl.expense_items:
            ws.cell(row=row, column=1, value=item.account_code)
            ws.cell(row=row, column=2, value=item.account_name)
            ws.cell(row=row, column=3, value=float(item.amount))
            row += 1
        ws.cell(row=row, column=2, value="Jami xarajat:").font = TOTAL_FONT
        ws.cell(row=row, column=3, value=float(pnl.total_expense)).font = TOTAL_FONT
        row += 2

        ws.cell(row=row, column=2, value="SOF FOYDA:").font = TOTAL_FONT
        ws.cell(row=row, column=3, value=float(pnl.net_profit)).font = TOTAL_FONT
        row += 1
        ws.cell(row=row, column=2, value="Foyda marjasi (%):").font = TOTAL_FONT
        ws.cell(row=row, column=3, value=float(round(pnl.profit_margin, 2)))

        self._auto_width(ws, [10, 40, 18])
        return self._to_bytes(wb)

    # ─── Yordamchi ────────────────────────────────────────────────────────────
    def _write_header(self, ws: Worksheet, headers: list[str], title: str = "") -> int:
        if title:
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
            header_row = 2
        else:
            header_row = 1

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        return header_row + 1

    def _auto_width(self, ws: Worksheet, widths: list[int]) -> None:
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _to_bytes(self, wb: Workbook) -> bytes:
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
