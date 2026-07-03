import io
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app.engines.report.pnl import PnLLineItem, ProfitAndLoss
from app.engines.report.trial_balance import TrialBalance, TrialBalanceLine
from app.models.all_models import Transaction, TransactionStatus, TransactionType
from app.services.document.excel import ExcelExportService


def _make_transaction(amount: str, description: str = "Test tranzaksiya") -> Transaction:
    return Transaction(
        transaction_date=date.today(),
        description=description,
        transaction_type=TransactionType.INCOME,
        total_amount=Decimal(amount),
        currency="UZS",
        status=TransactionStatus.CONFIRMED,
        ai_generated=False,
    )


class TestExportTransactions:
    def test_produces_loadable_xlsx_with_correct_rows(self):
        transactions = [_make_transaction("100000", "Sotuv 1"), _make_transaction("200000", "Sotuv 2")]
        service = ExcelExportService()

        xlsx_bytes = service.export_transactions(transactions, company_name="Test MCHJ")
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        assert ws.cell(row=1, column=1).value == "Test MCHJ"
        assert ws.cell(row=2, column=1).value == "Sana"
        assert ws.cell(row=3, column=2).value == "Sotuv 1"
        assert ws.cell(row=4, column=2).value == "Sotuv 2"

    def test_handles_empty_transaction_list(self):
        service = ExcelExportService()

        xlsx_bytes = service.export_transactions([], company_name="Test MCHJ")
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert wb.active is not None


class TestExportTrialBalance:
    def test_produces_loadable_xlsx_with_balance_status(self):
        tb = TrialBalance(
            company_id="test",
            period_start=None,
            period_end=None,
            lines=[
                TrialBalanceLine("1010", "Kassa", "asset", Decimal("100000"), Decimal("0"), Decimal("100000")),
                TrialBalanceLine("4010", "Sotuv", "income", Decimal("0"), Decimal("100000"), Decimal("100000")),
            ],
            total_debit=Decimal("100000"),
            total_credit=Decimal("100000"),
        )
        service = ExcelExportService()

        xlsx_bytes = service.export_trial_balance(tb, company_name="Test MCHJ")
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        assert ws.cell(row=3, column=1).value == "1010"


class TestExportPnl:
    def test_produces_loadable_xlsx_with_net_profit(self):
        pnl = ProfitAndLoss(
            company_id="test",
            period_start=None,
            period_end=None,
            income_items=[PnLLineItem("4010", "Sotuv daromadi", Decimal("500000"))],
            expense_items=[PnLLineItem("6010", "Ijara xarajati", Decimal("100000"))],
        )
        service = ExcelExportService()

        xlsx_bytes = service.export_pnl(pnl, company_name="Test MCHJ")
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert wb.active is not None
        found_net_profit_label = any(
            row[1].value == "SOF FOYDA:" for row in wb.active.iter_rows() if row[1].value
        )
        assert found_net_profit_label
